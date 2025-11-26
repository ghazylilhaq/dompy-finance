"""
Import Service - handles parsing and importing transaction files.
"""

import csv
import io
from collections import defaultdict
from datetime import datetime
from decimal import Decimal, InvalidOperation
from uuid import UUID

from sqlalchemy.orm import Session

from app.schemas.import_profile import (
    ParsedRow,
    ParseResult,
    MappingItem,
    ImportResult,
    PreviewRow,
    PreviewResult,
)
from app.schemas.transaction import TransactionCreate, TransferCreate
from app.crud import import_profile as import_crud
from app.crud import transaction as transaction_crud
from app.crud import category as category_crud
from app.crud import account as account_crud


# Expected column headers (case-insensitive)
EXPECTED_COLUMNS = {
    "id": ["id"],
    "date": ["date"],
    "categories": ["categories", "category"],
    "amount": ["amount"],
    "accounts": ["accounts", "account"],
    "description": ["description", "desc"],
}


def normalize_header(header: str) -> str:
    """Normalize a header string for matching."""
    return header.strip().lower()


def find_column_index(headers: list[str], column_key: str) -> int | None:
    """Find the index of a column by its key, checking alternative names."""
    normalized_headers = [normalize_header(h) for h in headers]
    for alt_name in EXPECTED_COLUMNS.get(column_key, []):
        if alt_name in normalized_headers:
            return normalized_headers.index(alt_name)
    return None


def parse_csv_content(content: bytes) -> tuple[list[str], list[list[str]]]:
    """
    Parse CSV content and return headers and rows.
    Tries different encodings.
    """
    for encoding in ["utf-8", "utf-8-sig", "latin-1", "cp1252"]:
        try:
            text = content.decode(encoding)
            reader = csv.reader(io.StringIO(text))
            rows = list(reader)
            if rows:
                return rows[0], rows[1:]
        except (UnicodeDecodeError, csv.Error):
            continue

    raise ValueError("Unable to parse CSV file. Please check the file encoding.")


def parse_excel_content(content: bytes) -> tuple[list[str], list[list[str]]]:
    """
    Parse Excel content and return headers and rows.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ValueError("Excel support requires openpyxl. Please install it.")

    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    if ws is None:
        raise ValueError("Excel file has no active worksheet.")

    rows = []
    for row in ws.iter_rows(values_only=True):
        # Convert None to empty string, ensure all values are strings
        rows.append([str(cell) if cell is not None else "" for cell in row])

    if not rows:
        raise ValueError("Excel file contains no data.")

    return rows[0], rows[1:]


def parse_file(file_content: bytes, filename: str) -> list[ParsedRow]:
    """
    Parse an import file and return a list of ParsedRow objects.

    Args:
        file_content: Raw file bytes
        filename: Original filename (used to detect format)

    Returns:
        List of ParsedRow objects

    Raises:
        ValueError: If file cannot be parsed or is missing required columns
    """
    filename_lower = filename.lower()

    if filename_lower.endswith(".csv"):
        headers, data_rows = parse_csv_content(file_content)
    elif filename_lower.endswith((".xlsx", ".xls")):
        headers, data_rows = parse_excel_content(file_content)
    else:
        raise ValueError(
            f"Unsupported file format. Please use CSV or Excel (.xlsx). Got: {filename}"
        )

    if not data_rows:
        raise ValueError("File contains no data rows.")

    # Find column indices
    col_indices = {}
    missing_columns = []

    for col_key in ["id", "date", "categories", "amount", "accounts", "description"]:
        idx = find_column_index(headers, col_key)
        if idx is None:
            missing_columns.append(col_key)
        else:
            col_indices[col_key] = idx

    if missing_columns:
        raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")

    # Parse each row
    parsed_rows = []
    for row_index, row in enumerate(data_rows):
        # Skip empty rows
        if not any(cell.strip() for cell in row):
            continue

        # Safely get values with bounds checking
        def get_value(key: str) -> str:
            idx = col_indices[key]
            if idx < len(row):
                return row[idx].strip()
            return ""

        external_id = get_value("id")
        date_str = get_value("date")
        category_value = get_value("categories")
        amount_str = get_value("amount")
        account_value = get_value("accounts")
        description = get_value("description")

        # Skip rows without essential data
        if not date_str or not amount_str:
            continue

        # Parse amount
        try:
            # Remove currency symbols and thousands separators
            amount_clean = (
                amount_str.replace(",", "").replace("$", "").replace("Rp", "").strip()
            )
            amount = Decimal(amount_clean)
        except (InvalidOperation, ValueError):
            continue  # Skip rows with invalid amounts

        parsed_rows.append(
            ParsedRow(
                row_index=row_index,
                external_id=external_id,
                date=date_str,
                category_value=category_value,
                account_value=account_value,
                amount=amount,
                description=description,
            )
        )

    if not parsed_rows:
        raise ValueError("No valid data rows found in file.")

    return parsed_rows


def analyze_mappings(
    db: Session, profile_id: UUID, parsed_rows: list[ParsedRow]
) -> ParseResult:
    """
    Analyze parsed rows and determine which values need mapping.

    Returns:
        ParseResult with unmapped values and existing mappings
    """
    # Collect distinct values
    category_values = set()
    account_values = set()

    for row in parsed_rows:
        if row.category_value:
            category_values.add(row.category_value)
        if row.account_value:
            account_values.add(row.account_value)

    # Get existing mappings
    existing_category_mappings = import_crud.get_value_mappings_dict(
        db, profile_id, "category"
    )
    existing_account_mappings = import_crud.get_value_mappings_dict(
        db, profile_id, "account"
    )

    # Convert UUIDs to strings for JSON serialization
    existing_cat_str = {k: str(v) for k, v in existing_category_mappings.items()}
    existing_acc_str = {k: str(v) for k, v in existing_account_mappings.items()}

    # Determine unmapped values
    unmapped_categories = [
        v for v in category_values if v not in existing_category_mappings
    ]
    unmapped_accounts = [
        v for v in account_values if v not in existing_account_mappings
    ]

    return ParseResult(
        profile_id=profile_id,
        total_rows=len(parsed_rows),
        parsed_rows=parsed_rows,
        unmapped_categories=sorted(unmapped_categories),
        unmapped_accounts=sorted(unmapped_accounts),
        existing_category_mappings=existing_cat_str,
        existing_account_mappings=existing_acc_str,
    )


def parse_date(date_str: str) -> datetime | None:
    """
    Parse a date string in various formats.
    Primary format: dd/MM/yyyy
    """
    formats = [
        "%d/%m/%Y",  # dd/MM/yyyy
        "%d-%m-%Y",  # dd-MM-yyyy
        "%Y-%m-%d",  # yyyy-MM-dd (ISO)
        "%m/%d/%Y",  # MM/dd/yyyy (US)
        "%d.%m.%Y",  # dd.MM.yyyy
    ]

    for fmt in formats:
        try:
            return datetime.strptime(date_str.strip(), fmt)
        except ValueError:
            continue

    return None


def get_transfer_category_ids(db: Session, user_id: str) -> dict[str, UUID] | None:
    """
    Get the transfer category IDs for a user.
    Returns {"incoming": UUID, "outgoing": UUID} or None if not found.

    Looks for categories by name at any level - works for both:
    - New users: root-level transfer categories
    - Existing users: transfer categories under "Transfer" parent (backward compatible)
    """
    # Find transfer categories by name (any level, for backward compatibility)
    incoming = category_crud.get_category_by_name(
        db, "Incoming transfer", user_id, any_parent=True
    )
    outgoing = category_crud.get_category_by_name(
        db, "Outgoing transfer", user_id, any_parent=True
    )

    if incoming and outgoing:
        return {"incoming": incoming.id, "outgoing": outgoing.id}

    return None


def detect_transfer_pairs(
    rows: list[ParsedRow],
    category_mappings: dict[str, UUID],
    transfer_category_ids: dict[str, UUID],
) -> tuple[list[tuple[ParsedRow, ParsedRow]], list[ParsedRow], list[str]]:
    """
    Detect transfer pairs from parsed rows.

    Returns:
        - transfer_pairs: List of (outgoing_row, incoming_row) tuples
        - regular_rows: Rows that are not part of a transfer pair
        - warnings: Any warnings about unmatched transfers
    """
    transfer_candidates = []
    regular_rows = []
    warnings = []

    incoming_id = transfer_category_ids["incoming"]
    outgoing_id = transfer_category_ids["outgoing"]

    # Separate rows into transfer candidates vs regular
    for row in rows:
        resolved_category_id = category_mappings.get(row.category_value)
        if resolved_category_id in (incoming_id, outgoing_id):
            transfer_candidates.append(row)
        else:
            regular_rows.append(row)

    if not transfer_candidates:
        return [], regular_rows, warnings

    # Group transfer candidates by (date, abs_amount) for matching
    groups: dict[tuple[str, Decimal], list[ParsedRow]] = defaultdict(list)
    for row in transfer_candidates:
        key = (row.date, abs(row.amount))
        groups[key].append(row)

    transfer_pairs = []
    unmatched = []

    for key, group_rows in groups.items():
        # Separate into outgoing (negative) and incoming (positive)
        outgoing_candidates = [r for r in group_rows if r.amount < 0]
        incoming_candidates = [r for r in group_rows if r.amount > 0]

        # Match pairs - ensure different accounts
        matched_outgoing = set()
        matched_incoming = set()

        for out_row in outgoing_candidates:
            out_account = out_row.account_value
            for in_row in incoming_candidates:
                in_account = in_row.account_value
                # Must be different accounts and not already matched
                if out_account != in_account and id(in_row) not in matched_incoming:
                    transfer_pairs.append((out_row, in_row))
                    matched_outgoing.add(id(out_row))
                    matched_incoming.add(id(in_row))
                    break

        # Collect unmatched
        for row in outgoing_candidates:
            if id(row) not in matched_outgoing:
                unmatched.append(row)
        for row in incoming_candidates:
            if id(row) not in matched_incoming:
                unmatched.append(row)

    # Add warnings for unmatched transfer rows
    for row in unmatched:
        warnings.append(
            f"Row {row.row_index + 1}: Transfer category but no matching pair found, "
            f"importing as regular transaction"
        )
        regular_rows.append(row)

    return transfer_pairs, regular_rows, warnings


def generate_preview(
    db: Session,
    user_id: str,
    profile_id: UUID,
    parsed_rows: list[ParsedRow],
    category_mappings: dict[str, UUID],
    account_mappings: dict[str, UUID],
) -> PreviewResult:
    """
    Generate a preview of the import showing resolved values and validation status.
    """
    # Get category and account details for name resolution
    all_categories = {c.id: c for c in category_crud.get_categories(db, user_id)}
    all_accounts = {a.id: a for a in account_crud.get_accounts(db, user_id)}

    # Get transfer category IDs (if they exist)
    transfer_cat_ids = get_transfer_category_ids(db, user_id)

    preview_rows: list[PreviewRow] = []
    warnings: list[str] = []

    # First pass: create preview rows with resolved values
    for row in parsed_rows:
        validation_errors = []

        # Parse date
        parsed_dt = parse_date(row.date)
        parsed_date_str = parsed_dt.isoformat() if parsed_dt else None
        if not parsed_dt:
            validation_errors.append(f"Invalid date format: {row.date}")

        # Resolve category
        category_id = category_mappings.get(row.category_value)
        category_name = None
        if category_id:
            if category_id in all_categories:
                category_name = all_categories[category_id].name
            else:
                # Mapping exists but category was deleted
                validation_errors.append(
                    f"Category no longer exists (update mapping for '{row.category_value}')"
                )
        elif row.category_value:
            validation_errors.append(f"No mapping for category: {row.category_value}")

        # Resolve account
        account_id = account_mappings.get(row.account_value)
        account_name = None
        if account_id:
            if account_id in all_accounts:
                account_name = all_accounts[account_id].name
            else:
                # Mapping exists but account was deleted
                validation_errors.append(
                    f"Account no longer exists (update mapping for '{row.account_value}')"
                )
        elif row.account_value:
            validation_errors.append(f"No mapping for account: {row.account_value}")

        # Determine type
        is_transfer = False
        if transfer_cat_ids and category_id in (
            transfer_cat_ids["incoming"],
            transfer_cat_ids["outgoing"],
        ):
            is_transfer = True
            tx_type = "transfer"
        elif row.amount < 0:
            tx_type = "expense"
        else:
            tx_type = "income"

        # Check for zero amount
        if row.amount == 0:
            validation_errors.append("Amount is zero")

        is_valid = len(validation_errors) == 0

        preview_rows.append(
            PreviewRow(
                row_index=row.row_index,
                external_id=row.external_id,
                date=row.date,
                parsed_date=parsed_date_str,
                amount=row.amount,
                type=tx_type,
                description=row.description,
                category_value=row.category_value,
                category_id=category_id,
                category_name=category_name,
                account_value=row.account_value,
                account_id=account_id,
                account_name=account_name,
                is_valid=is_valid,
                validation_errors=validation_errors,
                is_transfer=is_transfer,
                transfer_pair_index=None,  # Will be set in second pass
            )
        )

    # Second pass: detect transfer pairs and update preview rows
    total_transfers = 0
    if transfer_cat_ids:
        # Build a mapping from row_index to preview_row index
        row_index_map = {pr.row_index: i for i, pr in enumerate(preview_rows)}

        # Detect pairs using the original parsed rows
        transfer_pairs, _, pair_warnings = detect_transfer_pairs(
            parsed_rows, category_mappings, transfer_cat_ids
        )
        warnings.extend(pair_warnings)
        total_transfers = len(transfer_pairs)

        # Update preview rows with pair information
        for out_row, in_row in transfer_pairs:
            out_idx = row_index_map.get(out_row.row_index)
            in_idx = row_index_map.get(in_row.row_index)
            if out_idx is not None and in_idx is not None:
                preview_rows[out_idx].transfer_pair_index = in_row.row_index
                preview_rows[in_idx].transfer_pair_index = out_row.row_index

    # Count valid/invalid
    total_valid = sum(1 for pr in preview_rows if pr.is_valid)
    total_invalid = sum(1 for pr in preview_rows if not pr.is_valid)

    return PreviewResult(
        profile_id=profile_id,
        rows=preview_rows,
        total_valid=total_valid,
        total_invalid=total_invalid,
        total_transfers=total_transfers,
        warnings=warnings,
    )


def execute_import(
    db: Session,
    user_id: str,
    profile_id: UUID,
    parsed_rows: list[ParsedRow],
    category_mappings: list[MappingItem],
    account_mappings: list[MappingItem],
    excluded_indices: list[int] | None = None,
) -> ImportResult:
    """
    Execute the import: persist mappings and create transactions.
    Automatically detects and creates transfer pairs.

    Args:
        db: Database session
        user_id: User ID
        profile_id: Import profile ID
        parsed_rows: Rows to import
        category_mappings: New category mappings to persist
        account_mappings: New account mappings to persist
        excluded_indices: Row indices to skip

    Returns:
        ImportResult with counts and any errors
    """
    errors = []
    imported_count = 0
    skipped_count = 0
    transfer_count = 0
    excluded_set = set(excluded_indices or [])

    # Filter out excluded rows
    rows_to_import = [r for r in parsed_rows if r.row_index not in excluded_set]
    skipped_count += len(parsed_rows) - len(rows_to_import)

    # Persist new mappings
    if category_mappings:
        import_crud.create_value_mappings_batch(
            db, profile_id, category_mappings, "category"
        )
    if account_mappings:
        import_crud.create_value_mappings_batch(
            db, profile_id, account_mappings, "account"
        )

    # Build complete mapping lookups
    all_category_mappings = import_crud.get_value_mappings_dict(
        db, profile_id, "category"
    )
    all_account_mappings = import_crud.get_value_mappings_dict(
        db, profile_id, "account"
    )

    # Check for transfer categories and detect pairs
    transfer_cat_ids = get_transfer_category_ids(db, user_id)
    transfer_pairs = []
    regular_rows = rows_to_import

    if transfer_cat_ids:
        transfer_pairs, regular_rows, pair_warnings = detect_transfer_pairs(
            rows_to_import, all_category_mappings, transfer_cat_ids
        )
        errors.extend(pair_warnings)

    # Get valid category and account IDs for this user (for validation)
    valid_category_ids = {c.id for c in category_crud.get_categories(db, user_id)}
    valid_account_ids = {a.id for a in account_crud.get_accounts(db, user_id)}

    # Process transfer pairs first
    processed_transfer_rows = set()
    for out_row, in_row in transfer_pairs:
        try:
            # Parse dates
            out_date = parse_date(out_row.date)
            in_date = parse_date(in_row.date)
            if not out_date or not in_date:
                errors.append(
                    f"Rows {out_row.row_index + 1} & {in_row.row_index + 1}: Invalid date"
                )
                skipped_count += 2
                continue

            # Resolve accounts
            from_account_id = all_account_mappings.get(out_row.account_value)
            to_account_id = all_account_mappings.get(in_row.account_value)

            if not from_account_id or not to_account_id:
                errors.append(
                    f"Rows {out_row.row_index + 1} & {in_row.row_index + 1}: Missing account mapping"
                )
                skipped_count += 2
                continue

            # Validate accounts exist
            if from_account_id not in valid_account_ids:
                errors.append(
                    f"Rows {out_row.row_index + 1} & {in_row.row_index + 1}: Source account no longer exists"
                )
                skipped_count += 2
                continue
            if to_account_id not in valid_account_ids:
                errors.append(
                    f"Rows {out_row.row_index + 1} & {in_row.row_index + 1}: Destination account no longer exists"
                )
                skipped_count += 2
                continue

            # Create transfer
            transfer_data = TransferCreate(
                from_account_id=from_account_id,
                to_account_id=to_account_id,
                amount=abs(out_row.amount),
                date=out_date,
                description=out_row.description
                or in_row.description
                or "Imported transfer",
                hide_from_summary=True,
            )

            transaction_crud.create_transfer(db, transfer_data, user_id)
            transfer_count += 1
            processed_transfer_rows.add(out_row.row_index)
            processed_transfer_rows.add(in_row.row_index)

        except Exception as e:
            db.rollback()  # Rollback the failed transaction to allow subsequent operations
            errors.append(
                f"Rows {out_row.row_index + 1} & {in_row.row_index + 1}: {str(e)}"
            )
            skipped_count += 2

    # Process regular rows
    for row in regular_rows:
        # Skip if already processed as part of a transfer
        if row.row_index in processed_transfer_rows:
            continue

        try:
            # Parse date
            parsed_date = parse_date(row.date)
            if not parsed_date:
                errors.append(
                    f"Row {row.row_index + 1}: Invalid date format '{row.date}'"
                )
                skipped_count += 1
                continue

            # Resolve category
            category_id = all_category_mappings.get(row.category_value)
            if not category_id:
                errors.append(
                    f"Row {row.row_index + 1}: No mapping for category '{row.category_value}'"
                )
                skipped_count += 1
                continue

            # Validate category exists
            if category_id not in valid_category_ids:
                errors.append(
                    f"Row {row.row_index + 1}: Category no longer exists (was deleted?)"
                )
                skipped_count += 1
                continue

            # Resolve account
            account_id = all_account_mappings.get(row.account_value)
            if not account_id:
                errors.append(
                    f"Row {row.row_index + 1}: No mapping for account '{row.account_value}'"
                )
                skipped_count += 1
                continue

            # Validate account exists
            if account_id not in valid_account_ids:
                errors.append(
                    f"Row {row.row_index + 1}: Account no longer exists (was deleted?)"
                )
                skipped_count += 1
                continue

            # Determine transaction type from amount sign
            amount = row.amount
            if amount == 0:
                skipped_count += 1
                continue

            tx_type = "expense" if amount < 0 else "income"
            abs_amount = abs(amount)

            # Create transaction
            tx_data = TransactionCreate(
                date=parsed_date,
                type=tx_type,
                amount=abs_amount,
                category_id=category_id,
                account_id=account_id,
                description=row.description or f"Imported: {row.external_id}",
                tags=[],
            )

            transaction_crud.create_transaction(db, tx_data, user_id)
            imported_count += 1

        except Exception as e:
            db.rollback()  # Rollback the failed transaction to allow subsequent operations
            errors.append(f"Row {row.row_index + 1}: {str(e)}")
            skipped_count += 1

    return ImportResult(
        imported_count=imported_count,
        skipped_count=skipped_count,
        transfer_count=transfer_count,
        errors=errors[:50],  # Limit errors to first 50
    )
